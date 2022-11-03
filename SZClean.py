import os
import re
import requests
import dns.resolver
import openpyxl as xl
from openpyxl.styles import Alignment

os.environ['no_proxy'] = '*'
requests.packages.urllib3.disable_warnings()
domain_set = set()
ip_set = set()
url_set = set()


def isResolve(url):
    domain = url.split(":")[0].split("/")[0]
    r = re.match(r"^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$",domain)
    if r:
        # print("t"+url)
        return True
    try:
        # print(domain)
        dns.resolver.resolve(domain,"A")
        return True
    except:
        return False

def isAlive(url):
    ret = -1
    t_url = "http://"+ url
    if(isResolve(url)):
        cookie_headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/101.0.4951.41 Safari/537.36 Edg/101.0.1210.32'}
        try:
            r = requests.get("http://"+url, headers=cookie_headers, timeout=5)
            ret = r.status_code
            if r.status_code != 200:
                r = requests.get("https://"+url, headers=cookie_headers, verify=False, timeout=5)
                t_url = "https://"+url
                ret = r.status_code
        except:
            pass
    return ret, t_url


def urlClean(url):
    """
    处理路径,判断存活太浪费时间了,相当于我扫描了一遍,这种工作直接交给gowitness做
    """
    ret,t_url = isAlive(url)
    ret = str(ret)
    if ret.startswith("2"):#Todo打印出来
        # print(t_url)
        url_set.add(t_url+"\n")
        pass
    elif ret.startswith("3"):#Todo 跟随跳转
        pass
    elif ret.startswith("5"):#Todo hosts碰撞，改refer头，改手机header，改XFF
        pass
    else:# 已经关站，就省略
        pass

def _dealItem(item):
    # 根据格式分流ip和域名
    r = re.match(r"^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$",item)
    if r:
        ip_set.add(item+"\n")
    else:
        domain_set.add(item+"\n")
    # print(item)
    return True

def dealItem(item):
    """
    处理域名和ip的字符串
    list 列表型的ip处理  []  ['domain']
    str  字符型的ip处理 无需处理
    """
    item_list = item.replace("[","").replace("]","").replace("'","").split(",")
    if type(item_list) == list:
        for d in item_list:
            _dealItem(d.strip())
    else:
        _dealItem(item.strip())

def dealXlsx(file_path):
    """
    水泽输出的表
    没啥用的：
    备案反查顶级域名（需要手动看的,可能需要合并处理）
    爱企查（如果有邮箱，邮箱手机号是有价值的，额外域名）
    服务(可能有用，但很少，做一个汇集展示)
    github敏感信息爬虫(得手动看,价值不在xlsx里面)
    证书(被包含在子域名A记录中,有些信任证书的域是其他公司资产,舍弃)
    Host碰撞(目前没遇到有用的情况,暂不处理，舍弃)
    IP反查域名(舍弃,容易在cdn的ip查到一堆域名)

    动态链接和后台地址(会出现业务url,价值相对高)
    网络空间搜索引擎(url简单路径,http头+ip/域名+端口)
    存活网站标题(url简单路径,http头+ip/域名+端口)
    子域名A记录(获取域名和ip,要排除cdn的情况)
    相关域名和C段(获取域名和ip,但域名中可能会出现ip,需要处理)

    域名 -> 子域名 -> 真实IP -> 业务端口
    Doing:域名/ip/简单url/业务Url
    TODO:域名备案信息/邮箱/手机号/github项目/服务
    """
    wb = xl.load_workbook(file_path)

    # 域名和IP汇总处理
    ws1 = wb['子域名A记录']
    tmp1_set = set()
    for row in ws1.iter_rows(min_row=2, max_col=3, max_row=ws1.max_row):
        # row[0]:domain row[1]:ip_list row[2]:cdn
        tmp1_set.add(row[0].value)
        if "NOT" in row[2].value:
            # 不是CDN,添加ip
            tmp1_set.add(row[1].value)
 
    ws2 = wb['相关域名和C段']
    for row in ws2.iter_rows(min_row=2, max_col=3, max_row=ws1.max_row):
        # print(row, row[0].value, row[1].value)
        tmp1_set.add(row[1].value)
        tmp1_set.add(row[0].value)

    # url处理
    tmp2_set = set()
    ws3 = wb['存活网站标题']
    for row in ws3.iter_rows(min_row=2, max_col=3, max_row=ws3.max_row):
        # row[1]:url
        # print(row[0].value)
        tmp2_set.add(row[0].value)

    ws4 = wb['网络空间搜索引擎']
    for row in ws4.iter_rows(min_row=2, max_col=6, max_row=ws4.max_row):
        # row[1]:url  row[3]:ip
        # print(row[1].value, row[3].value)
        tmp2_set.add(row[1].value)
        tmp1_set.add(row[3].value)

    ws5 = wb['动态链接和后台地址']
    for row in ws5.iter_rows(min_row=2, max_col=2, max_row=ws5.max_row):
        # row[1]:url 会有一个中文标题
        if "http" in row[0].value:
            # print(row[0].value)
            tmp2_set.add(row[0].value)

    for u in tmp2_set:
        if u != None and u != "[]":
            u = u.strip().lstrip("https://").lstrip("http://")
            d = u.split(":")[0].split("/")[0]
            print(d,u)
            tmp1_set.add(d)
            url_set.add(u+"\n")

    for i in tmp1_set:
        if i != None and i != "[]":
            dealItem(i)

def save():
    """
    储存domain/ip/url三个资产
    """
    # ip和域名用于移交扫描器进行端口，资产探测
    with open("t.txt","w") as f:
        f.writelines(domain_set.union(ip_set))
        pass

    # url用于gowitness网站截图
    with open("url.txt","w") as ff:
        ff.writelines(url_set)

def main(path):
    root = os.path.abspath(path)
    for f in os.listdir(path):
        fp = os.path.join(root,f)
        if os.path.isfile(fp) and fp.endswith(".xlsx") and not f.startswith("."):
            print(f)
            dealXlsx(fp)
            break
    
    save()

if __name__ == "__main__":
    path = input("输入结果文件夹：")
    if(os.path.isdir(path)):
        main(path)
    else:
        print("路径错误")